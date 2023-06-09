import openpyxl

wb = openpyxl.load_workbook(r'MAGICMANDIVIDEOA01.xlsx')
mNG = []
mCh = []
mNG_2 = {}
mCh_2 = {}
sheetNames = wb.sheetnames
for i in wb.sheetnames:
    maxRow = wb[i].max_row
    temp = [i]
    benchMark = wb[i].cell(3, 3).value
    for row in range(4, maxRow+1):
        wb[i].cell(row, 4).value = wb[i].cell(row, 3).value - benchMark
    for row in range(4, maxRow+1, 2):
        if wb[i].cell(row, 4).value <= 0 or wb[i].cell(row + 1, 4).value <= 0:
            wb[i].cell(row, 5).value = 'Error'
            temp.append('Error')
        else:
            wb[i].cell(row, 5).value = wb[i].cell(row, 4).value / wb[i].cell(row + 1, 4).value
            temp.append(wb[i].cell(row, 4).value / wb[i].cell(row + 1, 4).value)
    c = 0
    s = 0
    for ele in temp[1:]:
        if ele != 'Error':
            s += ele
            c += 1
    A = round(s / c, 2)
    temp.append(A)
    temp.append(benchMark)
    if 'mNG' in i:
        mNG.append(temp)
        mNG_2[i[1:3]] = A
    else:
        mCh.append(temp)
        mCh_2[i[1:3]] = A

res3 = wb.create_sheet('mNG_vs_mCh', 0)
mCh_res1 = wb.create_sheet('mCh_res1', 0)
mNG_res1 = wb.create_sheet('mNG_res1', 0)

A = {}
for i in range(len(mNG)):
    temp = 0
    c = 0
    for row in range(1, len(mNG[i])):
        wb['mNG_res1'].cell(row, i+2).value = mNG[i][row-1]
        wb['mCh_res1'].cell(row, i+2).value = mCh[i][row-1]
        if row == 1:
            wb['mNG_vs_mCh'].cell(row, i + 2).value = mNG[i][row-1] + '/' + mCh[i][row-1]
        elif row == len(mNG[i]) - 1:
            wb['mNG_vs_mCh'].cell(row, i + 2).value = round(temp/c, 2)
            A[mNG[i][0][1:4]] = round(temp/c, 2)
        else:
            if mNG[i][row-1] == 'Error' or mCh[i][row-1] == 'Error':
                wb['mNG_vs_mCh'].cell(row, i + 2).value = 'Error'
            else:
                wb['mNG_vs_mCh'].cell(row, i+2).value = mNG[i][row-1]/mCh[i][row-1]
                temp += mNG[i][row-1]/mCh[i][row-1]
                c += 1
e1 = 0
e2 = 0
maxRow = wb['mNG_res1'].max_row
for i in range(len(mNG)):
    wb['mNG_res1'].cell(maxRow+1, i+2).value = mNG[i][-1]
    wb['mCh_res1'].cell(maxRow+1, i+2).value = mCh[i][-1]
    wb['mNG_res1'].cell(maxRow+2, i + 2).value = mNG[i].count('Error')
    e1 += mNG[i].count('Error')
    wb['mCh_res1'].cell(maxRow+2, i + 2).value = mCh[i].count('Error')
    e2 += mCh[i].count('Error')

wb['mNG_res1'].cell(maxRow, 1).value = 'Average'
wb['mCh_res1'].cell(maxRow, 1).value = 'Average'
wb['mNG_res1'].cell(maxRow+1, 1).value = 'Benchmark'
wb['mCh_res1'].cell(maxRow+1, 1).value = 'Benchmark'
wb['mNG_res1'].cell(maxRow+2, 1).value = 'Error'
wb['mCh_res1'].cell(maxRow+2, 1).value = 'Error'
wb['mNG_res1'].cell(maxRow+3, 1).value = 'Total Error'
wb['mCh_res1'].cell(maxRow+3, 1).value = 'Total Error'
wb['mNG_res1'].cell(maxRow+3, 2).value = e1
wb['mCh_res1'].cell(maxRow+3, 2).value = e2
wb['mNG_vs_mCh'].cell(maxRow, 1).value = 'Average'


wb.save('res.xlsx')





